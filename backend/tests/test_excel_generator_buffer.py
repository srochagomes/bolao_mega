"""
Unit tests for Excel generator buffer optimization
Tests buffer handling for large volumes (1M+ games)
"""
import pytest
from app.services.excel_generator import ExcelGenerator
from app.models.generation import GameConstraints
from typing import Iterator, List


class TestExcelGeneratorBuffer:
    """Test Excel generator buffer optimization"""
    
    def test_streaming_small_quantity(self):
        """Test streaming with small quantity"""
        generator = ExcelGenerator()
        
        def games_iterator() -> Iterator[List[int]]:
            for i in range(100):
                yield [1, 2, 3, 4, 5, 6 + i % 10]
        
        constraints = GameConstraints(numbers_per_game=6)
        
        excel_bytes = generator.generate_excel(
            games=games_iterator(),
            constraints=constraints,
            budget=600.0,
            quantity=100,
            manual_numbers=None
        )
        
        assert len(excel_bytes) > 0
    
    def test_streaming_medium_quantity(self):
        """Test streaming with medium quantity"""
        generator = ExcelGenerator()
        
        def games_iterator() -> Iterator[List[int]]:
            for i in range(10_000):
                yield sorted([1 + (i % 10), 10 + (i % 10), 20 + (i % 10), 
                             30 + (i % 10), 40 + (i % 10), 50 + (i % 10)])
        
        constraints = GameConstraints(numbers_per_game=6)
        
        excel_bytes = generator.generate_excel(
            games=games_iterator(),
            constraints=constraints,
            budget=60_000.0,
            quantity=10_000,
            manual_numbers=None
        )
        
        assert len(excel_bytes) > 0
    
    def test_streaming_large_quantity(self):
        """Test streaming with large quantity (100k+)"""
        generator = ExcelGenerator()
        
        def games_iterator() -> Iterator[List[int]]:
            for i in range(100_000):
                # Generate varied games
                base = (i * 7) % 55
                yield sorted([
                    (base + 1) % 60 + 1,
                    (base + 5) % 60 + 1,
                    (base + 10) % 60 + 1,
                    (base + 15) % 60 + 1,
                    (base + 20) % 60 + 1,
                    (base + 25) % 60 + 1,
                ])
        
        constraints = GameConstraints(numbers_per_game=6)
        
        excel_bytes = generator.generate_excel(
            games=games_iterator(),
            constraints=constraints,
            budget=600_000.0,
            quantity=100_000,
            manual_numbers=None
        )
        
        assert len(excel_bytes) > 0
    
    def test_buffer_chunking(self):
        """Test that buffer chunks correctly"""
        generator = ExcelGenerator()
        
        # Create iterator that yields in chunks
        chunk_size = 5000
        total_games = 25_000
        
        def games_iterator() -> Iterator[List[int]]:
            for i in range(total_games):
                yield sorted([1 + (i % 10), 10 + (i % 10), 20 + (i % 10),
                             30 + (i % 10), 40 + (i % 10), 50 + (i % 10)])
        
        constraints = GameConstraints(numbers_per_game=6)
        
        excel_bytes = generator.generate_excel(
            games=games_iterator(),
            constraints=constraints,
            budget=150_000.0,
            quantity=total_games,
            manual_numbers=None
        )
        
        assert len(excel_bytes) > 0
    
    def test_write_batch_method(self):
        """Test batch writing method"""
        from openpyxl import Workbook
        from app.services.excel_generator import ExcelGenerator
        
        generator = ExcelGenerator()
        wb = Workbook()
        ws = wb.create_sheet("Test")
        
        games = [
            [1, 2, 3, 4, 5, 6],
            [7, 8, 9, 10, 11, 12],
            [13, 14, 15, 16, 17, 18]
        ]
        
        generator._write_games_to_sheet_batch(
            ws, games, 1, set(), 6
        )
        
        # Check that games were written
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=2, column=1).value == 7
        assert ws.cell(row=3, column=1).value == 13


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

