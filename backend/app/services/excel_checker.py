"""
Excel file checker service
Checks games in Excel file against drawn numbers
Supports checking multiple split files transparently
"""
from openpyxl import load_workbook
from typing import List, Dict, Optional
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelChecker:
    """Service to check Excel files against drawn numbers"""
    
    def check_file(self, file_content: bytes, drawn_numbers: List[int]) -> Dict:
        """
        Check an Excel file against drawn numbers
        Returns count of quadras (4), quinas (5), and senas (6)
        
        Args:
            file_content: Excel file content as bytes
            drawn_numbers: List of 6 drawn numbers
            
        Returns:
            Dictionary with counts of quadras, quinas, and senas
        """
        drawn_set = set(drawn_numbers)
        
        if len(drawn_set) != 6:
            raise ValueError("Drawn numbers must be 6 unique numbers")
        
        # Load workbook from bytes
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        
        # Try to find the "Generated Games" sheet
        games_sheet = None
        for sheet_name in workbook.sheetnames:
            if "Jogos Gerados" in sheet_name or "Generated Games" in sheet_name or "games" in sheet_name.lower():
                games_sheet = workbook[sheet_name]
                break
        
        if not games_sheet:
            # Try to find any sheet that might contain games
            # Usually it's the second sheet (index 1)
            if len(workbook.sheetnames) > 1:
                games_sheet = workbook[workbook.sheetnames[1]]
            else:
                games_sheet = workbook[workbook.sheetnames[0]]
        
        # Count matches
        quadras = 0
        quinas = 0
        senas = 0
        total_games = 0
        
        # Find the start row (usually row 4, after headers)
        start_row = 4
        max_row = games_sheet.max_row
        
        # Check each row for games
        for row_idx in range(start_row, max_row + 1):
            game_numbers = []
            
            # Read numbers from columns (support variable number of columns)
            for col_idx in range(1, 20):  # Support up to 17 numbers per game
                cell_value = games_sheet.cell(row=row_idx, column=col_idx).value
                
                # Stop if we hit an empty cell or non-numeric value
                if cell_value is None:
                    break
                
                try:
                    number = int(cell_value)
                    if 1 <= number <= 60:
                        game_numbers.append(number)
                    else:
                        break
                except (ValueError, TypeError):
                    break
            
            # Process if we have at least 6 numbers (support games with more numbers)
            if len(game_numbers) >= 6:
                total_games += 1
                # Count matches
                matches = len(set(game_numbers) & drawn_set)
                
                if matches == 4:
                    quadras += 1
                elif matches == 5:
                    quinas += 1
                elif matches == 6:
                    senas += 1
        
        return {
            "quadras": quadras,
            "quinas": quinas,
            "senas": senas,
            "total_games_checked": total_games
        }
    
    def check_file_by_path(self, file_path: Path, drawn_numbers: List[int]) -> Dict:
        """
        Check an Excel file by file path
        """
        with open(file_path, 'rb') as f:
            file_content = f.read()
        return self.check_file(file_content, drawn_numbers)
    
    def check_multiple_files(
        self,
        file_contents: List[bytes],
        drawn_numbers: List[int]
    ) -> Dict:
        """
        Check multiple Excel files and aggregate results
        Used for split files - checks all files transparently
        
        Args:
            file_contents: List of Excel file contents as bytes
            drawn_numbers: List of 6 drawn numbers
            
        Returns:
            Dictionary with aggregated counts of quadras, quinas, and senas
        """
        total_quadras = 0
        total_quinas = 0
        total_senas = 0
        total_games = 0
        
        logger.info(f"Checking {len(file_contents)} files against drawn numbers {drawn_numbers}")
        
        for idx, file_content in enumerate(file_contents, 1):
            try:
                result = self.check_file(file_content, drawn_numbers)
                total_quadras += result['quadras']
                total_quinas += result['quinas']
                total_senas += result['senas']
                total_games += result['total_games_checked']
                logger.debug(
                    f"File {idx}/{len(file_contents)}: "
                    f"{result['total_games_checked']} games, "
                    f"{result['quadras']} quadras, {result['quinas']} quinas, {result['senas']} senas"
                )
            except Exception as e:
                logger.error(f"Error checking file {idx}: {e}", exc_info=True)
                # Continue with other files even if one fails
        
        logger.info(
            f"Total results: {total_games} games checked, "
            f"{total_quadras} quadras, {total_quinas} quinas, {total_senas} senas"
        )
        
        return {
            "quadras": total_quadras,
            "quinas": total_quinas,
            "senas": total_senas,
            "total_games_checked": total_games,
            "files_checked": len(file_contents)
        }


# Singleton instance
excel_checker = ExcelChecker()

