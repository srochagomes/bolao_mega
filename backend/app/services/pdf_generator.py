"""
PDF generator for Mega-Sena lottery tickets
Generates HTML first, then converts to PDF using WeasyPrint
Based on official Mega-Sena ticket format from lotoloto.com.br
Format: Landscape A4, 3 volantes per page, 3 games per volante
"""
from weasyprint import HTML, CSS
from openpyxl import load_workbook
from typing import List, Set
import io
import logging
import random
import string

logger = logging.getLogger(__name__)


class PDFGenerator:
    """Generates PDF tickets for Mega-Sena lottery"""
    
    GAMES_PER_TICKET = 3
    TICKETS_PER_PAGE = 3
    NUMBERS_PER_ROW = 10
    NUMBERS_PER_COL = 6
    
    def generate_pdf(self, excel_file_path: str) -> bytes:
        """
        Generate PDF from Excel file
        Returns PDF bytes
        """
        # Load Excel and extract games
        games = self._extract_games_from_excel(excel_file_path)
        
        if not games:
            raise ValueError("No games found in Excel file")
        
        # Generate HTML
        html_content = self._generate_html(games)
        
        # Debug: save HTML to file for inspection
        import os
        debug_path = '/tmp/debug_volantes.html'
        with open(debug_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logger.info(f"Debug HTML saved to {debug_path} ({len(html_content)} chars)")
        
        # Convert HTML to PDF with explicit DPI
        html_doc = HTML(string=html_content)
        # Use 96 DPI (standard web DPI) to match pixel sizes
        pdf_bytes = html_doc.write_pdf()
        
        return pdf_bytes
    
    def generate_html_file(self, excel_file_path: str) -> str:
        """
        Generate HTML file for printing
        Returns HTML content as string
        """
        # Load Excel and extract games
        games = self._extract_games_from_excel(excel_file_path)
        
        if not games:
            raise ValueError("No games found in Excel file")
        
        # Generate HTML
        html_content = self._generate_html(games)
        
        return html_content
    
    def _extract_games_from_excel(self, excel_file_path: str) -> List[List[int]]:
        """Extract games from Excel file"""
        workbook = load_workbook(excel_file_path, data_only=True)
        
        # Find the "Generated Games" sheet
        games_sheet = None
        for sheet_name in workbook.sheetnames:
            if "Generated Games" in sheet_name or "games" in sheet_name.lower():
                games_sheet = workbook[sheet_name]
                break
        
        if not games_sheet:
            if len(workbook.sheetnames) > 1:
                games_sheet = workbook[workbook.sheetnames[1]]
            else:
                games_sheet = workbook[workbook.sheetnames[0]]
        
        games = []
        start_row = 4
        max_row = games_sheet.max_row
        
        logger.info(f"Reading games from rows {start_row} to {max_row}")
        
        for row_idx in range(start_row, max_row + 1):
            game_numbers = []
            has_data = False
            
            # Read all 6 columns
            for col_idx in range(1, 7):
                cell_value = games_sheet.cell(row=row_idx, column=col_idx).value
                
                if cell_value is not None:
                    has_data = True
                    try:
                        number = int(cell_value)
                        if 1 <= number <= 60:
                            game_numbers.append(number)
                        else:
                            # Invalid number, skip this row
                            break
                    except (ValueError, TypeError):
                        # Not a number, skip this row
                        break
            
            # Only add if we have exactly 6 numbers
            if len(game_numbers) == 6:
                sorted_game = sorted(game_numbers)
                games.append(sorted_game)
                logger.debug(f"Row {row_idx}: Added game {sorted_game}")
            elif has_data and len(game_numbers) > 0:
                # Partial game found, log warning
                logger.warning(f"Row {row_idx} has incomplete game: {len(game_numbers)} numbers")
        
        logger.info(f"Extracted {len(games)} complete games from Excel file")
        if len(games) > 0:
            logger.info(f"First game: {games[0]}, Last game: {games[-1]}")
        return games
    
    def _generate_ticket_code(self) -> str:
        """Generate a random 5-character code like official tickets"""
        return ''.join(random.choices(string.ascii_uppercase, k=5))
    
    def _generate_html(self, games: List[List[int]]) -> str:
        """Generate HTML for tickets"""
        logger.info(f"Generating HTML for {len(games)} games")
        
        # Group games into tickets (3 games per ticket)
        tickets = []
        for i in range(0, len(games), self.GAMES_PER_TICKET):
            ticket_games = games[i:i + self.GAMES_PER_TICKET]
            tickets.append(ticket_games)
        
        logger.info(f"Created {len(tickets)} tickets from {len(games)} games")
        
        # Log ticket distribution
        total_games_in_tickets = sum(len(t) for t in tickets)
        logger.info(f"Total games in tickets: {total_games_in_tickets} (expected {len(games)})")
        if total_games_in_tickets != len(games):
            logger.error(f"MISMATCH: {total_games_in_tickets} games in tickets vs {len(games)} games extracted!")
        
        html_parts = []
        html_parts.append(self._get_html_header())
        
        # Generate pages with 3 tickets each (landscape)
        for page_idx in range(0, len(tickets), self.TICKETS_PER_PAGE):
            page_tickets = tickets[page_idx:page_idx + self.TICKETS_PER_PAGE]
            page_num = page_idx // self.TICKETS_PER_PAGE + 1
            logger.info(f"Generating page {page_num} with {len(page_tickets)} tickets")
            
            html_parts.append('<div class="page-container">')
            for ticket_idx, ticket_games in enumerate(page_tickets):
                ticket_num = page_idx + ticket_idx + 1
                logger.info(f"  Adding ticket {ticket_num}/{len(tickets)} with {len(ticket_games)} games: {ticket_games}")
                ticket_html = self._generate_ticket_html(ticket_num, len(tickets), ticket_games)
                html_parts.append(ticket_html)
            html_parts.append('</div>')
            logger.info(f"  Closed page-container for page {page_num}")
        
        html_parts.append("</body></html>")
        
        return "\n".join(html_parts)
    
    def _get_html_header(self) -> str:
        """Get HTML header with CSS - using exact pixel sizes from official CSS"""
        return """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Volantes Mega-Sena</title>
<style>
body, html { margin: 0; padding: 0; font-family: Tahoma, Arial, sans-serif; }
.previaImpressao { padding: 0; margin: 0; width: 100%; }
.page-container { 
  page-break-after: always; 
  page-break-inside: avoid;
  width: 297mm; 
  height: 210mm;
  display: block;
  overflow: visible;
  position: relative;
  box-sizing: border-box;
  margin: 0;
  padding: 0;
  min-height: 210mm;
}
.page-container:last-child { page-break-after: auto; }
.page-container::after {
  content: "";
  display: table;
  clear: both;
  height: 0;
  line-height: 0;
  visibility: hidden;
}
.volante { 
  position: relative; 
  z-index: 1400; 
  padding: 0; 
  margin: 0; 
  margin-left: 8px; 
  margin-top: 0; 
  height: 506px; 
  width: 310px; 
  float: left; 
  page-break-inside: avoid;
  display: block;
  box-sizing: border-box;
}
.cabecalho { font-family: Tahoma, Arial, sans-serif; font-size: 10px; text-align: center; position: relative; top: 11px; height: 140px; padding-top: 4px; width: 317px; margin-left: 3px; box-sizing: border-box; overflow: visible; }
.cabecalho .jogos { font-size: 8px; line-height: 1.2; }
.cabecalho .jogo { padding: 2px 0; white-space: nowrap; }
.cabecalho > div { padding: 2px 0; }
.cabecalho > .composicao { text-decoration: underline; }
.cabecalho > .loteria { font-weight: 700; font-size: 1.2em; }
.cabecalho > .site { display: none; }
.entorno { 
  position: relative; 
  width: 325px; 
  margin-left: 15px; 
  margin-top: 11px; 
  margin-bottom: 0;
  margin-right: 0;
  float: left; 
  page-break-after: avoid; 
  page-break-inside: avoid;
  page-break-before: avoid;
  display: block;
  clear: none;
  box-sizing: border-box;
  height: auto;
  overflow: visible;
  visibility: visible;
  opacity: 1;
}
.entorno.borda_vermelha { 
  border-left: none;
  border-top: none;
  border-bottom: none;
  border-right: dashed 1px #fcc; 
  box-sizing: border-box; 
}
.entorno > .aviso { display: none; }
.previaImpressao.megasena.normal .qd { width: 12px; height: 5px; margin-right: 11.5px; position: relative; top: 1px; border-radius: 2px; overflow: hidden; float: left; }
.previaImpressao.megasena.normal .qd.ch { background-color: #000; }
.previaImpressao.megasena.normal .qd > img { width: 12px; height: 5px; position: absolute; top: 0; left: 0; opacity: 0; }
.previaImpressao.megasena.normal .qd.ch > img { opacity: 1; }
.previaImpressao.megasena.normal .linha { height: 7px; margin-bottom: 5px; clear: both; }
.previaImpressao.megasena.normal .apostas { position: relative; margin-top: 30px; }
.previaImpressao.megasena.normal .aposta { margin-left: 41px; margin-bottom: 28px; }
.previaImpressao.megasena.normal .aposta:nth-child(2) { margin-bottom: 28px; }
.previaImpressao.megasena.normal .aposta:nth-child(3) { margin-bottom: 27px; }
.previaImpressao.megasena.normal .aposta:nth-child(4) { margin-bottom: 0; }
.previaImpressao.megasena.normal .volante .codigos { position: absolute; top: -15px; left: 0; }
.previaImpressao.megasena.normal .volante .codigos > .horizontal { position: absolute; top: -2px; left: 0; width: 302px; margin-left: 7.5px; }
.previaImpressao.megasena.normal .volante .codigos > .horizontal > .blocoH { width: 15px; height: 6px; margin-left: 8.5px; background-color: #000; float: left; }
.previaImpressao.megasena.normal .volante .codigos > .horizontal > .blocoH.vazio { opacity: 0; }
.previaImpressao.megasena.normal .volante .codigos > .vertical { position: absolute; top: -7px; left: 0; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area { position: relative; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area2 { margin-top: 11px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area3 { margin-top: 17px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area4 { margin-top: 17px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area5 { margin-top: 22px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area6 { margin-top: 29px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area7 { margin-top: 31px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .area8 { margin-top: 26px; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .blocoV { width: 12px; height: 7px; margin-top: 5px; background-color: #000; position: relative; }
.previaImpressao.megasena.normal .volante .codigos > .vertical .blocoVP { width: 12px; height: 6px; margin-top: 5px; background-color: #000; position: relative; }
.previaImpressao.megasena.normal .linha_quantos_numeros { margin-left: 41px; margin-top: 33px; }
.previaImpressao.megasena.normal .linha_quantos_numeros .linha .qd { margin-right: 11.5px; }
.previaImpressao.megasena.normal .linha_surpresinha { margin-left: 41px; margin-top: 16px; height: 8px; }
.previaImpressao.megasena.normal .linha_teimosinha { margin-left: 41px; margin-top: 30px; height: 8px; }
.previaImpressao.megasena.normal .linha_bolao { margin-left: 41px; margin-top: 26px; }
.previaImpressao.megasena.normal .linha_bolao .linha { height: 7px; margin-bottom: 5px; }
.previaImpressao.megasena.normal .linha_bolao .linha > .qd { margin-right: 11.5px; }
@media print {
  @page { 
    size: A4 landscape; 
    margin: 0;
  }
  body {
    margin: 0;
    padding: 0;
  }
}
@page { 
  size: A4 landscape; 
  margin: 0;
  size: 297mm 210mm;
}
.entorno {
  page-break-before: avoid;
  page-break-after: avoid;
  page-break-inside: avoid;
}
.page-container .entorno:first-child {
  page-break-before: auto;
}
body { 
  width: 297mm; 
  height: 210mm;
  margin: 0;
  padding: 0;
}
</style>
</head>
<body>
<div class="previaImpressao megasena normal">"""
    
    def _generate_ticket_html(self, ticket_num: int, total_tickets: int, games: List[List[int]]) -> str:
        """Generate HTML for a single ticket with correct machine codes"""
        code = self._generate_ticket_code()
        # Calculate first game number based on ticket number (each ticket has 3 games)
        first_game = (ticket_num - 1) * self.GAMES_PER_TICKET + 1
        last_game = first_game + len(games) - 1
        
        html = [f'<div class="entorno borda_vermelha">']
        
        # Header
        html.append('<div class="cabecalho">')
        html.append('<div class="loteria">Mega-Sena</div>')
        html.append(f'<div class="composicao">{code}</div>')
        html.append(f'<div class="titulo">Volante {ticket_num} de {total_tickets}</div>')
        html.append(f'<div class="subtitulo">Jogos {first_game} a {last_game}</div>')
        html.append('<div class="jogos">')
        for idx, game in enumerate(games):
            game_str = ", ".join([f"{num:02d}" for num in game])
            html.append(f'<div class="jogo">Jogo {first_game + idx}: {game_str}</div>')
        html.append('</div>')
        html.append('</div>')
        
        # Volante with codes and games
        html.append('<div class="volante">')
        html.append('<div class="codigos">')
        
        # Horizontal codes (columns) - based on ALL games in this ticket
        html.append('<div class="horizontal">')
        all_marked_cols = set()
        for game in games:
            for num in game:
                col = (num - 1) % self.NUMBERS_PER_ROW
                all_marked_cols.add(col)
        
        for col in range(self.NUMBERS_PER_ROW):
            if col in all_marked_cols:
                html.append('<div class="blocoH"><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
            else:
                html.append('<div class="blocoH vazio"></div>')
        html.append('</div>')
        
        # Vertical codes (rows) - ALL areas must ALWAYS be present in FIXED positions
        # The areas are positional markers for the machine reader - they must be in the same
        # position on every ticket, regardless of which numbers are marked
        html.append('<div class="vertical">')
        
        # Mega-Sena uses areas 2-8 (no area1)
        # Map rows (0-5) to areas (2-7), and area8 is for quantity marker
        # ALL areas must ALWAYS be present in the SAME ORDER for fixed positioning
        row_to_area = {
            0: 'area2',  # First row (1-10) -> area2
            1: 'area3',  # Second row (11-20) -> area3
            2: 'area4',  # Third row (21-30) -> area4
            3: 'area5',  # Fourth row (31-40) -> area5
            4: 'area6',  # Fifth row (41-50) -> area6
            5: 'area7',  # Sixth row (51-60) -> area7
        }
        
        # Generate ALL areas (2-7) - ALWAYS present in FIXED order, even if empty
        # The CSS margin-top positions each area at a fixed vertical position
        for row in range(self.NUMBERS_PER_COL):
            area_class = row_to_area[row]
            html.append(f'<div class="area {area_class}">')
            # Count how many games have numbers in this row
            games_with_row = [game for game in games if any((num - 1) // self.NUMBERS_PER_ROW == row for num in game)]
            # One blocoV per game that has numbers in this row
            # If no games have numbers in this row, the area is empty but still present
            for _ in games_with_row:
                html.append(f'<div class="blocoV"><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
            html.append('</div>')
        
        # Area8 is for quantity marker (6 dezenas) - ALWAYS present with blocoVP for each game
        html.append('<div class="area area8">')
        for _ in range(len(games)):
            html.append(f'<div class="blocoVP"><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
        html.append('</div>')
        
        html.append('</div>')
        html.append('</div>')
        
        # Games (apostas) - each game has its own grid
        html.append('<div class="apostas">')
        for game_idx, game in enumerate(games):
            html.append('<div class="aposta">')
            game_set = set(game)
            
            # Draw 6 lines (rows) of 10 numbers each
            for row in range(self.NUMBERS_PER_COL):
                html.append('<div class="linha">')
                for col in range(self.NUMBERS_PER_ROW):
                    number = row * self.NUMBERS_PER_ROW + col + 1
                    if number in game_set:
                        html.append('<div class="qd ch"><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
                    else:
                        html.append('<div class="qd"></div>')
                html.append('</div>')
            html.append('</div>')
        
        html.append('</div>')  # Close apostas
        
        # Quantidade de números (6 dezenas for Mega-Sena)
        html.append('<div class="linha_quantos_numeros">')
        html.append('<div class="linha">')
        # Mark 6 (Mega-Sena always has 6 numbers)
        html.append('<div class="qd ch"><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
        for _ in range(5):
            html.append('<div class="qd"></div>')
        html.append('</div>')
        html.append('</div>')
        
        # Surpresinha, teimosinha, bolão (empty for now)
        html.append('<div class="linha_surpresinha"></div>')
        html.append('<div class="linha_teimosinha"></div>')
        html.append('<div class="linha_bolao">')
        html.append('<div class="linha_bolao_dezena"><div class="linha">')
        for _ in range(9):
            html.append('<div class="qd"></div>')
        html.append('</div></div>')
        html.append('<div class="linha_bolao_unidade"><div class="linha">')
        for _ in range(10):
            html.append('<div class="qd"></div>')
        html.append('</div></div>')
        html.append('<div class="linha_bolao_centena"><div class="linha">')
        html.append('<div class="qd "><img src="data:image/gif;base64,R0lGODlhAQABAIAAAAUEBAAAACwAAAAAAQABAAACAkQBADs="></div>')
        html.append('</div></div>')
        html.append('</div>')
        
        html.append('</div>')
        html.append('</div>')
        
        return "\n".join(html)

# Singleton instance
pdf_generator = PDFGenerator()
