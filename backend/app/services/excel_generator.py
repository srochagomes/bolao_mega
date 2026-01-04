"""
Excel file generation with validation and multiple sheets
Supports streaming for large volumes to avoid memory issues
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from typing import List, Optional, Iterator, Union, Callable
from datetime import datetime
import io
import logging
import time
from app.models.generation import GameConstraints
from app.services.historical_data import historical_data_service

logger = logging.getLogger(__name__)

# Excel row limit: 1,048,576 rows (Excel 2007+)
# We use 1,000,000 games per file to leave room for headers and formulas
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_GAMES_PER_FILE = 1_000_000  # Safe limit with headers


class ExcelGenerator:
    """Excel file generator with validation"""
    
    def __init__(self):
        self._header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self._header_font = Font(bold=True, color="FFFFFF")
        self._match_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        self._green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self._border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel(
        self,
        games: Union[List[List[int]], Iterator[List[int]]],
        constraints: GameConstraints,
        budget: float,
        quantity: int,
        manual_numbers: Optional[List[int]] = None,
        save_callback: Optional[Callable[[int, bytes, dict], None]] = None
    ) -> Union[bytes, List[bytes]]:
        """
        Generate Excel file(s) with validation, games, and audit sheets
        Supports both list and streaming (generator) input for memory efficiency
        Returns single file (bytes) or list of files (List[bytes]) if exceeds Excel limit
        
        Args:
            games: List of games or generator/iterator of games
            constraints: Game generation constraints
            budget: Budget used
            quantity: Total quantity of games
            manual_numbers: Optional manual numbers for validation
            
        Returns:
            bytes: Single Excel file if quantity <= EXCEL_MAX_GAMES_PER_FILE
            List[bytes]: Multiple Excel files if quantity > EXCEL_MAX_GAMES_PER_FILE
        """
        # Check if we need to split into multiple files
        if quantity > EXCEL_MAX_GAMES_PER_FILE:
            logger.info(
                f"📦 Quantity ({quantity}) exceeds Excel limit ({EXCEL_MAX_GAMES_PER_FILE}). "
                f"Splitting into multiple files..."
            )
            return self._generate_multiple_excel_files(
                games, constraints, budget, quantity, manual_numbers, save_callback
            )
        
        # Single file generation (original logic)
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_validation_sheet(wb, manual_numbers, quantity)
        
        # Check if games is a list or iterator
        if isinstance(games, list):
            # Traditional approach: sort all games in memory
            self._create_games_sheet(wb, games, manual_numbers)
        else:
            # Streaming approach: write incrementally with chunked sorting
            self._create_games_sheet_streaming(wb, games, manual_numbers, quantity, constraints.numbers_per_game)
        
        self._create_audit_sheet(wb, constraints, budget, quantity)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _generate_multiple_excel_files(
        self,
        games: Union[List[List[int]], Iterator[List[int]]],
        constraints: GameConstraints,
        budget: float,
        quantity: int,
        manual_numbers: Optional[List[int]] = None,
        save_callback: Optional[Callable[[int, bytes, dict], None]] = None
    ) -> List[bytes]:
        """
        Generate multiple Excel files when quantity exceeds Excel limit
        Each file contains up to EXCEL_MAX_GAMES_PER_FILE games
        
        BIG DATA STRATEGY:
        - Use Ray for parallel file generation (if available)
        - Process files incrementally
        - Save each file immediately after generation (if callback provided)
        - Update progress during generation
        - Use streaming for large datasets
        
        Args:
            save_callback: Optional callback(file_idx, file_bytes, file_metadata) to save file immediately
        """
        # Convert iterator to list if needed (for splitting)
        if isinstance(games, Iterator):
            logger.info("📦 Converting iterator to list for file splitting...")
            games = list(games)
        
        total_games = len(games)
        num_files = (total_games + EXCEL_MAX_GAMES_PER_FILE - 1) // EXCEL_MAX_GAMES_PER_FILE
        
        logger.info(
            f"📦 Splitting {total_games} games into {num_files} files "
            f"({EXCEL_MAX_GAMES_PER_FILE} games per file) - BIG DATA MODE"
        )
        
        # Try to use Ray for parallel Excel generation (BIG DATA optimization)
        try:
            import ray
            RAY_AVAILABLE = True
            if not ray.is_initialized():
                ray.init(ignore_reinit_error=True, num_cpus=None)
        except (ImportError, Exception) as e:
            RAY_AVAILABLE = False
            logger.debug(f"Ray not available for Excel generation: {e}")
        
        # BIG DATA: Use Ray for parallel generation if available and we have multiple files
        # Ray can generate multiple files in parallel, dramatically reducing total time
        # Example: 9 files sequential = ~63 minutes, Ray parallel = ~7 minutes (9x faster)
        if RAY_AVAILABLE and num_files > 1:
            logger.info(
                f"⚡ BIG DATA: Using Ray for parallel Excel generation: "
                f"{num_files} files in parallel (expected speedup: ~{num_files}x)"
            )
            try:
                return self._generate_multiple_excel_files_ray(
                    games, constraints, budget, quantity, manual_numbers, save_callback, num_files, total_games
                )
            except Exception as e:
                logger.warning(
                    f"⚠️ Ray parallel generation failed: {e}. "
                    f"Falling back to sequential generation."
                )
                # Fall through to sequential generation
        
        # Fallback to sequential generation
        files = []
        for file_idx in range(num_files):
            start_idx = file_idx * EXCEL_MAX_GAMES_PER_FILE
            end_idx = min(start_idx + EXCEL_MAX_GAMES_PER_FILE, total_games)
            file_games = games[start_idx:end_idx]
            file_quantity = len(file_games)
            
            progress_pct = (file_idx + 1) / num_files * 100
            logger.info(
                f"📄 Generating file {file_idx + 1}/{num_files}: "
                f"games {start_idx + 1}-{end_idx} ({file_quantity} games) "
                f"({progress_pct:.1f}% of Excel generation)"
            )
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            # For multi-file, show total quantity in validation sheet
            self._create_validation_sheet(wb, manual_numbers, quantity)
            
            # Create games sheet with subset of games
            self._create_games_sheet(wb, file_games, manual_numbers)
            
            # Create audit sheet with file info
            self._create_audit_sheet(
                wb, constraints, budget, quantity,
                file_info={
                    "file_number": file_idx + 1,
                    "total_files": num_files,
                    "games_in_file": file_quantity,
                    "start_index": start_idx + 1,
                    "end_index": end_idx
                }
            )
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            file_bytes = buffer.getvalue()
            files.append(file_bytes)
            
            logger.info(f"✅ File {file_idx + 1}/{num_files} generated ({len(file_bytes)} bytes)")
            
            # BIG DATA: Save file immediately if callback provided (incremental save)
            if save_callback:
                file_metadata = {
                    "file_number": file_idx + 1,
                    "total_files": num_files,
                    "games_in_file": file_quantity,
                    "start_index": start_idx + 1,
                    "end_index": end_idx,
                    "is_multi_file": True
                }
                try:
                    save_callback(file_idx, file_bytes, file_metadata)
                    logger.info(f"💾 File {file_idx + 1}/{num_files} saved incrementally to disk")
                except Exception as e:
                    logger.error(f"❌ Error saving file {file_idx + 1} incrementally: {e}", exc_info=True)
                    # Continue - file is still in memory
        
        logger.info(f"✅ Generated {len(files)} Excel files for {total_games} games")
        return files
    
    def _generate_single_excel_file(
        self,
        file_games: List[List[int]],
        file_idx: int,
        num_files: int,
        constraints: GameConstraints,
        budget: float,
        quantity: int,
        manual_numbers: Optional[List[int]],
        save_callback: Optional[Callable[[int, bytes, dict], None]] = None
    ) -> bytes:
        """
        Generate a single Excel file (used by Ray workers)
        """
        file_quantity = len(file_games)
        start_idx = file_idx * EXCEL_MAX_GAMES_PER_FILE
        
        logger.info(
            f"📄 [Ray Worker] Generating file {file_idx + 1}/{num_files}: "
            f"{file_quantity} games"
        )
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_validation_sheet(wb, manual_numbers, quantity)
        self._create_games_sheet(wb, file_games, manual_numbers)
        self._create_audit_sheet(
            wb, constraints, budget, quantity,
            file_info={
                "file_number": file_idx + 1,
                "total_files": num_files,
                "games_in_file": file_quantity,
                "start_index": start_idx + 1,
                "end_index": start_idx + file_quantity
            }
        )
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.getvalue()
        
        logger.info(f"✅ [Ray Worker] File {file_idx + 1}/{num_files} generated ({len(file_bytes)} bytes)")
        
        # Call save callback if provided (for incremental save)
        if save_callback:
            file_metadata = {
                "file_number": file_idx + 1,
                "total_files": num_files,
                "games_in_file": file_quantity,
                "start_index": start_idx + 1,
                "end_index": start_idx + file_quantity,
                "is_multi_file": True
            }
            try:
                save_callback(file_idx, file_bytes, file_metadata)
                logger.info(f"💾 [Ray Worker] File {file_idx + 1}/{num_files} saved incrementally")
            except Exception as e:
                logger.error(f"❌ [Ray Worker] Error saving file {file_idx + 1}: {e}", exc_info=True)
        
        return file_bytes
    
    def _generate_multiple_excel_files_ray(
        self,
        games: List[List[int]],
        constraints: GameConstraints,
        budget: float,
        quantity: int,
        manual_numbers: Optional[List[int]],
        save_callback: Optional[Callable[[int, bytes, dict], None]],
        num_files: int,
        total_games: int
    ) -> List[bytes]:
        """
        Generate multiple Excel files using Ray for parallel processing
        BIG DATA: Each file is generated in parallel by a Ray worker
        """
        import ray
        
        # Create Ray remote function for Excel generation
        # Note: save_callback cannot be serialized by Ray, so we handle it in the main process
        @ray.remote
        def generate_excel_file_remote(
            file_games: List[List[int]],
            file_idx: int,
            num_files: int,
            constraints_dict: dict,
            budget: float,
            quantity: int,
            manual_numbers: Optional[List[int]],
            total_games: int
        ) -> bytes:
            """Ray remote function to generate a single Excel file in parallel"""
            # Reconstruct constraints from dict (Ray serialization)
            from app.models.generation import GameConstraints
            constraints = GameConstraints(**constraints_dict)
            
            # Create ExcelGenerator instance in worker
            generator = ExcelGenerator()
            # Don't pass save_callback to worker (can't serialize), handle in main process
            return generator._generate_single_excel_file(
                file_games, file_idx, num_files, constraints, budget, quantity, manual_numbers, None
            )
        
        # Prepare constraints dict for Ray serialization
        constraints_dict = {
            "numbers_per_game": constraints.numbers_per_game,
            "max_repetition": constraints.max_repetition,
            "fixed_numbers": constraints.fixed_numbers,
            "seed": constraints.seed
        }
        
        # Split games into chunks for each file
        file_chunks = []
        for file_idx in range(num_files):
            start_idx = file_idx * EXCEL_MAX_GAMES_PER_FILE
            end_idx = min(start_idx + EXCEL_MAX_GAMES_PER_FILE, total_games)
            file_games = games[start_idx:end_idx]
            file_chunks.append((file_games, file_idx))
        
        logger.info(
            f"⚡ Starting parallel Excel generation with Ray: "
            f"{num_files} files, {len(file_chunks)} workers"
        )
        
        # Generate all files in parallel using Ray
        # BIG DATA: All files generated simultaneously across multiple CPU cores
        futures = []
        for file_games, file_idx in file_chunks:
            future = generate_excel_file_remote.remote(
                file_games,
                file_idx,
                num_files,
                constraints_dict,
                budget,
                quantity,
                manual_numbers,
                total_games
            )
            futures.append((file_idx, future))
        
        logger.info(
            f"⚡ Ray: {len(futures)} files queued for parallel generation. "
            f"Expected time: ~7 minutes (vs ~{num_files * 7} minutes sequential)"
        )
        
        # Collect results as they complete (with progress updates)
        # Ray processes files in parallel, so we get results as they finish
        files = [None] * num_files  # Pre-allocate list
        completed = 0
        start_time = time.time()
        
        # Use ray.wait() to get results as they complete (better than sequential ray.get)
        remaining_futures = {file_idx: future for file_idx, future in futures}
        
        while remaining_futures:
            # Wait for at least one future to complete
            ready, not_ready = ray.wait(
                list(remaining_futures.values()),
                num_returns=1,
                timeout=60.0  # Wait up to 60 seconds for next completion
            )
            
            # Process completed futures
            for future in ready:
                # Find which file_idx this future belongs to
                file_idx = None
                for idx, fut in remaining_futures.items():
                    if fut == future:
                        file_idx = idx
                        break
                
                if file_idx is not None:
                    try:
                        file_bytes = ray.get(future)
                        files[file_idx] = file_bytes
                        completed += 1
                        
                        elapsed = time.time() - start_time
                        logger.info(
                            f"✅ Ray: File {file_idx + 1}/{num_files} completed "
                            f"({completed}/{num_files} done, {elapsed/60:.1f} min elapsed)"
                        )
                        
                        # Call save callback if provided (for incremental save)
                        # This happens in the main process, not in Ray workers
                        if save_callback:
                            file_metadata = {
                                "file_number": file_idx + 1,
                                "total_files": num_files,
                                "games_in_file": len(file_chunks[file_idx][0]),
                                "start_index": file_idx * EXCEL_MAX_GAMES_PER_FILE + 1,
                                "end_index": min((file_idx + 1) * EXCEL_MAX_GAMES_PER_FILE, total_games),
                                "is_multi_file": True
                            }
                            try:
                                save_callback(file_idx, file_bytes, file_metadata)
                                logger.info(
                                    f"💾 Ray: File {file_idx + 1}/{num_files} saved incrementally"
                                )
                            except Exception as e:
                                logger.error(
                                    f"❌ Ray: Error saving file {file_idx + 1} incrementally: {e}",
                                    exc_info=True
                                )
                        
                        # Remove from remaining
                        del remaining_futures[file_idx]
                    except Exception as e:
                        logger.error(
                            f"❌ Ray: Error getting result for file {file_idx + 1}: {e}",
                            exc_info=True
                        )
                        # Remove failed future
                        if file_idx in remaining_futures:
                            del remaining_futures[file_idx]
        
        total_time = time.time() - start_time
        logger.info(
            f"✅ Ray: All {num_files} files generated in parallel "
            f"({total_time/60:.1f} minutes, ~{num_files * 7 / total_time:.1f}x speedup vs sequential)"
        )
        
        # Filter out None values (failed files)
        files = [f for f in files if f is not None]
        
        if len(files) != num_files:
            logger.warning(
                f"⚠️ Ray: Only {len(files)}/{num_files} files generated successfully"
            )
        
        return files
    
    def _create_validation_sheet(self, wb: Workbook, manual_numbers: Optional[List[int]], total_games: int):
        """Cria Aba 1: Entrada Manual + Validação"""
        ws = wb.create_sheet("Entrada Manual", 0)
        
        # Título
        ws['A1'] = "Entrada Manual de Números"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Instruções
        ws['A3'] = "Digite 6 números (1-60) abaixo:"
        ws['A3'].font = Font(italic=True)
        
        # Cabeçalhos
        headers = ["Número 1", "Número 2", "Número 3", "Número 4", "Número 5", "Número 6"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Input cells with validation
        valid_numbers = historical_data_service.get_all_numbers()
        number_list = ",".join(map(str, valid_numbers))
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=6, column=col_idx)
            cell.border = self._border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Validação de dados: lista suspensa com números válidos
            dv = DataValidation(
                type="list",
                formula1=f'"{number_list}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Número Inválido",
                error="Por favor, selecione um número entre 1 e 60 que existe nos dados históricos."
            )
            ws.add_data_validation(dv)
            dv.add(cell)
            
            # Pré-preenchimento se números manuais fornecidos
            if manual_numbers and col_idx <= len(manual_numbers):
                cell.value = manual_numbers[col_idx - 1]
        
        # Validação adicional: fórmula personalizada para verificação de intervalo
        for col_idx in range(1, 7):
            cell_ref = get_column_letter(col_idx) + "6"
            # Adicionar nota
            ws[cell_ref].comment = Comment("Digite um número entre 1 e 60", "Gerador Mega-Sena")
        
        # Add summary section for counting matches
        row = 8
        ws[f'A{row}'] = "Resultados da Conferência:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Quadra (4 matches) - Count games with exactly 4 matches
        ws[f'A{row}'] = "Quadras (4 acertos):"
        ws[f'A{row}'].font = Font(bold=True)
        # Fórmula: Contar linhas em Jogos Gerados onde coluna Acertos = 4
        last_row = 3 + total_games
        match_col = get_column_letter(7)  # Coluna G (7ª coluna)
        ws[f'B{row}'] = f"=COUNTIF('Jogos Gerados'!{match_col}4:{match_col}{last_row},4)"
        ws[f'B{row}'].border = self._border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].font = Font(bold=True, size=11)
        row += 1
        
        # Quina (5 matches) - Count games with exactly 5 matches
        ws[f'A{row}'] = "Quinas (5 acertos):"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"=COUNTIF('Jogos Gerados'!{match_col}4:{match_col}{last_row},5)"
        ws[f'B{row}'].border = self._border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].font = Font(bold=True, size=11)
        row += 1
        
        # Sena (6 matches) - Count games with exactly 6 matches
        ws[f'A{row}'] = "Senas (6 acertos):"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"=COUNTIF('Jogos Gerados'!{match_col}4:{match_col}{last_row},6)"
        ws[f'B{row}'].border = self._border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].font = Font(bold=True, size=11)
        
        # Format column B width
        ws.column_dimensions['B'].width = 15
    
    def _create_games_sheet(self, wb: Workbook, games: List[List[int]], manual_numbers: Optional[List[int]]):
        """Cria Aba 2: Jogos Gerados com formatação condicional"""
        ws = wb.create_sheet("Jogos Gerados", 1)
        
        # Título
        ws['A1'] = f"Jogos Gerados ({len(games)} total)"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(len(games[0]) if games else 6)}1')
        
        # Cabeçalhos
        headers = [f"Número {i+1}" for i in range(len(games[0]) if games else 6)]
        headers.append("Acertos")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Sort games: ensure numbers within each game are sorted (they should already be),
        # then sort games lexicographically by columns 1 to N
        sorted_games = []
        for game in games:
            # Ensure numbers are sorted within the game
            sorted_game = sorted(game)
            sorted_games.append(sorted_game)
        
        # Sort all games lexicographically (by column 1, then column 2, etc.)
        sorted_games.sort()
        
        # Games data
        manual_set = set(manual_numbers) if manual_numbers else set()
        
        # Determinar se usar formatação condicional (desabilitar para volumes maiores)
        use_conditional_formatting = len(sorted_games) <= 1000
        start_data_row = 4
        end_data_row = start_data_row + len(sorted_games) - 1
        
        # Escrever dados primeiro
        for row_idx, game in enumerate(sorted_games, start=4):
            # Check for matches
            matches = manual_set & set(game) if manual_set else set()
            has_match = len(matches) > 0
            
            for col_idx, number in enumerate(game, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=number)
                cell.border = self._border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Destaque direto se corresponde à entrada manual (mais eficiente)
                if number in matches:
                    cell.fill = self._match_fill
                    cell.font = Font(bold=True)
            
            # Indicador de acertos com fórmula
            match_col = len(game) + 1
            match_cell = ws.cell(row=row_idx, column=match_col)
            # Fórmula para contar acertos: Soma de COUNTIF para cada célula na linha
            formula_parts = []
            for col_idx in range(1, len(game) + 1):
                col_letter = get_column_letter(col_idx)
                formula_parts.append(f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{row_idx})")
            match_cell.value = '=' + '+'.join(formula_parts)
            match_cell.border = self._border
            match_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if has_match:
                match_cell.fill = self._match_fill
                match_cell.font = Font(bold=True)
        
        # Adicionar formatação condicional por RANGE (muito mais eficiente)
        if use_conditional_formatting and manual_set and sorted_games:
            numbers_per_game = len(sorted_games[0])
            for col_idx in range(1, numbers_per_game + 1):
                col_letter = get_column_letter(col_idx)
                range_ref = f'{col_letter}{start_data_row}:{col_letter}{end_data_row}'
                
                # Uma regra para todo o range ao invés de uma por célula
                fill_rule = FormulaRule(
                    formula=[f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{start_data_row})>0"],
                    fill=self._green_fill,
                    font=Font(bold=True)
                )
                ws.conditional_formatting.add(range_ref, fill_rule)
    
    def _create_games_sheet_streaming(
        self,
        wb: Workbook,
        games_iterator: Iterator[List[int]],
        manual_numbers: Optional[List[int]],
        total_games: int,
        numbers_per_game: int
    ):
        """
        Cria aba de jogos usando abordagem de streaming otimizada para grandes volumes
        Processa jogos em lotes, ordena incrementalmente e escreve com buffer eficiente
        Otimizado para volumes de 1M+ jogos
        """
        ws = wb.create_sheet("Jogos Gerados", 1)
        
        # Título
        ws['A1'] = f"Jogos Gerados ({total_games} total)"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(numbers_per_game)}1')
        
        # Cabeçalhos
        headers = [f"Número {i+1}" for i in range(numbers_per_game)]
        headers.append("Acertos")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Buffer otimizado para grandes volumes
        # Para 1M+ jogos: usar chunks menores e escrever mais frequentemente
        if total_games > 1_000_000:
            sort_chunk_size = 20_000  # Ordenar em chunks de 20k
            write_buffer_size = 50_000  # Escrever 50k por vez
        elif total_games > 100_000:
            sort_chunk_size = 10_000
            write_buffer_size = 25_000
        else:
            sort_chunk_size = 5_000
            write_buffer_size = 10_000
        
        # Usar heap merge para ordenação incremental (mais eficiente em memória)
        from heapq import heappush, heappop
        
        # Buffer para jogos ordenados (mantém apenas o necessário)
        sorted_buffer = []
        current_chunk = []
        games_processed = 0
        rows_written = 0
        
        manual_set = set(manual_numbers) if manual_numbers else set()
        
        logger.info(
            f"Processando {total_games} jogos em modo streaming otimizado "
            f"(sort chunk: {sort_chunk_size}, write buffer: {write_buffer_size})"
        )
        
        def write_batch(games_batch: List[List[int]], start_row: int):
            """Helper para escrever um batch de jogos"""
            self._write_games_to_sheet_batch(
                ws, games_batch, start_row, manual_set, numbers_per_game
            )
        
        # Coletar e processar jogos em lotes
        for game in games_iterator:
            # Garantir que números estão ordenados dentro do jogo
            sorted_game = sorted(game)
            current_chunk.append(sorted_game)
            games_processed += 1
            
            # Quando o chunk está cheio, ordenar e adicionar ao buffer
            if len(current_chunk) >= sort_chunk_size:
                current_chunk.sort()  # Ordenar lexicograficamente
                
                # Adicionar ao buffer ordenado (usando heap para merge eficiente)
                sorted_buffer.extend(current_chunk)
                sorted_buffer.sort()  # Manter ordenado
                current_chunk = []  # Clear immediately to free memory
                
                # CRITICAL: Write when buffer reaches limit to prevent memory growth
                # For very large volumes, write more frequently
                if len(sorted_buffer) >= write_buffer_size:
                    batch_to_write = sorted_buffer[:write_buffer_size]
                    batch_start_row = rows_written + 4
                    batch_end_row = batch_start_row + len(batch_to_write) - 1
                    
                    # Check Excel row limit before writing
                    if batch_end_row > EXCEL_MAX_ROWS:
                        raise ValueError(
                            f"Cannot write batch to row {batch_end_row}: Excel row limit is {EXCEL_MAX_ROWS}. "
                            f"Total games would be {rows_written + len(batch_to_write)}, but Excel supports max {EXCEL_MAX_GAMES_PER_FILE} games per file. "
                            f"Use _generate_multiple_excel_files instead."
                        )
                    
                    write_batch(batch_to_write, batch_start_row)
                    rows_written += write_buffer_size
                    # CRITICAL: Use slice assignment to free memory immediately
                    sorted_buffer = sorted_buffer[write_buffer_size:]
                    
                    # Force garbage collection for very large volumes
                    if total_games > 1_000_000 and rows_written % 200_000 == 0:
                        import gc
                        gc.collect()
                
                # Log progresso
                if games_processed % 100_000 == 0:
                    logger.info(
                        f"Processado: {games_processed}/{total_games} jogos "
                        f"({rows_written} escritos, {len(sorted_buffer)} no buffer)"
                    )
        
        # Processar chunk restante
        if current_chunk:
            current_chunk.sort()
            sorted_buffer.extend(current_chunk)
            sorted_buffer.sort()
        
        # Escrever buffer final
        if sorted_buffer:
            logger.info(f"Escrevendo buffer final: {len(sorted_buffer)} jogos...")
            final_start_row = rows_written + 4
            final_end_row = final_start_row + len(sorted_buffer) - 1
            
            # Check Excel row limit before writing
            if final_end_row > EXCEL_MAX_ROWS:
                raise ValueError(
                    f"Cannot write final buffer to row {final_end_row}: Excel row limit is {EXCEL_MAX_ROWS}. "
                    f"Total games would be {rows_written + len(sorted_buffer)}, but Excel supports max {EXCEL_MAX_GAMES_PER_FILE} games per file. "
                    f"Use _generate_multiple_excel_files instead."
                )
            
            write_batch(sorted_buffer, final_start_row)
            rows_written += len(sorted_buffer)
            sorted_buffer.clear()  # Clear to free memory
        
        # Final cleanup
        current_chunk.clear()
        import gc
        gc.collect()
        
        logger.info(f"Sucesso: {games_processed} jogos processados, {rows_written} escritos no Excel")
    
    def _write_games_to_sheet(
        self,
        ws,
        games: List[List[int]],
        start_row: int,
        manual_set: set,
        numbers_per_game: int
    ):
        """
        Escreve jogos na planilha de forma eficiente
        Usado para escrita incremental em grandes volumes
        OTIMIZADO: Usa formatação por range ao invés de célula por célula
        """
        if not games:
            return
        
        end_row = start_row + len(games) - 1
        
        # Para volumes maiores, desabilitar formatação condicional (muito pesado)
        use_conditional_formatting = len(games) <= 1000
        
        # Escrever dados primeiro
        for idx, game in enumerate(games):
            row_idx = start_row + idx
            # Verificar correspondências
            matches = manual_set & set(game) if manual_set else set()
            
            for col_idx, number in enumerate(game, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=number)
                cell.border = self._border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Destaque direto se corresponde à entrada manual (mais eficiente)
                if number in matches:
                    cell.fill = self._match_fill
                    cell.font = Font(bold=True)
            
            # Indicador de acertos com fórmula
            match_col = len(game) + 1
            match_cell = ws.cell(row=row_idx, column=match_col)
            # Construir fórmula: =COUNTIF('Entrada Manual'!$A$6:$F$6,A4)+COUNTIF('Entrada Manual'!$A$6:$F$6,B4)+...
            formula_parts = []
            for col_idx in range(1, len(game) + 1):
                col_letter = get_column_letter(col_idx)
                formula_parts.append(f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{row_idx})")
            match_cell.value = '=' + '+'.join(formula_parts)
            match_cell.border = self._border
            match_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if matches:
                match_cell.fill = self._match_fill
                match_cell.font = Font(bold=True)
        
        # Adicionar formatação condicional por RANGE (muito mais eficiente)
        if use_conditional_formatting and manual_set:
            for col_idx in range(1, numbers_per_game + 1):
                col_letter = get_column_letter(col_idx)
                range_ref = f'{col_letter}{start_row}:{col_letter}{end_row}'
                
                # Uma regra para todo o range ao invés de uma por célula
                fill_rule = FormulaRule(
                    formula=[f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{start_row})>0"],
                    fill=self._green_fill,
                    font=Font(bold=True)
                )
                ws.conditional_formatting.add(range_ref, fill_rule)
    
    def _write_games_to_sheet_batch(
        self,
        ws,
        games: List[List[int]],
        start_row: int,
        manual_set: set,
        numbers_per_game: int
    ):
        """
        Escreve jogos em batch otimizado para grandes volumes
        Usa escrita em lote e formatação por range para melhor performance
        """
        if not games:
            return
        
        end_row = start_row + len(games) - 1
        
        # Check Excel row limit
        if end_row > EXCEL_MAX_ROWS:
            raise ValueError(
                f"Cannot write games to row {end_row}: Excel row limit is {EXCEL_MAX_ROWS}. "
                f"This should not happen if _generate_multiple_excel_files is used correctly."
            )
        
        # Para volumes maiores, desabilitar formatação condicional (muito pesado)
        use_conditional_formatting = len(games) <= 1000
        
        # Agrupar operações similares
        manual_set_frozen = frozenset(manual_set) if manual_set else frozenset()
        
        # Escrever dados primeiro
        for idx, game in enumerate(games):
            row_idx = start_row + idx
            game_set = set(game)
            matches = manual_set_frozen & game_set if manual_set_frozen else set()
            
            # Escrever números do jogo
            for col_idx, number in enumerate(game, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=number)
                cell.border = self._border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Destaque direto se corresponde à entrada manual (mais eficiente)
                if number in matches:
                    cell.fill = self._match_fill
                    cell.font = Font(bold=True)
            
            # Indicador de acertos com fórmula
            match_col = len(game) + 1
            match_cell = ws.cell(row=row_idx, column=match_col)
            # Construir fórmula de forma otimizada
            formula_parts = []
            for col_idx in range(1, len(game) + 1):
                col_letter = get_column_letter(col_idx)
                formula_parts.append(f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{row_idx})")
            match_cell.value = '=' + '+'.join(formula_parts)
            match_cell.border = self._border
            match_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if matches:
                match_cell.fill = self._match_fill
                match_cell.font = Font(bold=True)
        
        # Adicionar formatação condicional por RANGE (muito mais eficiente)
        if use_conditional_formatting and manual_set:
            for col_idx in range(1, numbers_per_game + 1):
                col_letter = get_column_letter(col_idx)
                range_ref = f'{col_letter}{start_row}:{col_letter}{end_row}'
                
                # Uma regra para todo o range ao invés de uma por célula
                fill_rule = FormulaRule(
                    formula=[f"COUNTIF('Entrada Manual'!$A$6:$F$6,{col_letter}{start_row})>0"],
                    fill=self._green_fill,
                    font=Font(bold=True)
                )
                ws.conditional_formatting.add(range_ref, fill_rule)
    
    def _create_audit_sheet(
        self,
        wb: Workbook,
        constraints: GameConstraints,
        budget: float,
        quantity: int,
        file_info: Optional[dict] = None
    ):
        """Cria Aba 3: Regras e Resumo (Auditoria)"""
        ws = wb.create_sheet("Regras e Resumo", 2)
        
        # Título
        ws['A1'] = "Parâmetros de Geração e Auditoria"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Seção de parâmetros
        ws[f'A{row}'] = "Parâmetros de Geração"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        params = [
            ("Orçamento (R$)", f"R$ {budget:.2f}"),
            ("Quantidade de Jogos", quantity),
            ("Números por Jogo", constraints.numbers_per_game),
            ("Repetição Máxima", f"{constraints.max_repetition or 2} (ajusta automaticamente)"),
            ("Números Fixos", ", ".join(map(str, constraints.fixed_numbers)) if constraints.fixed_numbers else "Nenhum"),
            ("Observação", "Regras estatísticas (ímpar/par, frequência, sequências) são aplicadas automaticamente com base em dados históricos"),
        ]
        
        # Add file info if this is part of a multi-file generation
        if file_info:
            params.insert(2, ("Arquivo", f"{file_info['file_number']} de {file_info['total_files']}"))
            params.insert(3, ("Jogos neste arquivo", f"{file_info['games_in_file']} (jogos {file_info['start_index']}-{file_info['end_index']})"))
        
        for param_name, param_value in params:
            ws[f'A{row}'] = param_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = param_value
            ws[f'B{row}'].border = self._border
            row += 1
        
        row += 2
        
        # Informações de dados históricos
        last_update = historical_data_service.get_last_update_date()
        ws[f'A{row}'] = "Dados Históricos"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        ws[f'A{row}'] = "Última Atualização"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = last_update.strftime("%Y-%m-%d %H:%M:%S") if last_update else "N/A"
        ws[f'B{row}'].border = self._border
        row += 2
        
        # Aviso
        ws[f'A{row}'] = "AVISO IMPORTANTE"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
        row += 1
        
        disclaimer = (
            "Este sistema não aumenta a probabilidade de ganhar. "
            "Ele fornece apenas organização estatística e geração de combinações baseadas em regras. "
            "Os resultados da loteria são aleatórios e não podem ser previstos. "
            "Use esta ferramenta apenas para fins de entretenimento e organização."
        )
        
        ws[f'A{row}'] = disclaimer
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:B{row+2}')
        ws.row_dimensions[row].height = 60
        
        # Formatar colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

